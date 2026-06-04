# excel_utils.py
'''
Создание Excel-базы кабельных журналов
Версионирование, перенос старых записей, обработка новых журналов
'''

import json
import openpyxl
import re
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import date
from docx import Document
from tqdm import tqdm

from doc_utils import take_all_docx_from_dir, extract_all_text_from_docx
from cable_parser import row_parser
import config
from config import regular_KKS_building, regular_cable_tray, regular_journal_kks, regular_journal_kks_short

def current_version_finder(dir_in, b_name='Cable base ver.'):
    """
    Находит последнюю версию кабельной базы в указанной папке.
    Args:
        dir_in: путь к папке с базами
        b_name: шаблон имени базы (по умолчанию 'Cable base ver.')
    Returns:
        int: номер следующей версии (текущая версия + 1)
    """
    current_version = 1
    local_list_of_bases_vers = []
    xlsx_files = list(Path(dir_in).glob(b_name + '*.xlsx'))
    
    for file in xlsx_files:
        dot_index = str(file.stem).rfind('.')
        local_list_of_bases_vers.append(int(str(file.stem)[dot_index + 1:]))
    
    if not local_list_of_bases_vers:
        print(f'Кабельные базы не найдены в указанной папке')
    else:
        current_version = max(local_list_of_bases_vers) + 1
        print(f'Найдена последняя версия базы {b_name}{current_version - 1}, текущая версия - {current_version}')
    
    return current_version


def remove_plus_from_numbers(obj):
    """
    Рекурсивно обходит списки и удаляет символ '+' из строк,
    которые начинаются с '+' и затем содержат число (целое или дробное).
    """
    if isinstance(obj, list):
        return [remove_plus_from_numbers(item) for item in obj]
    elif isinstance(obj, str):
        if obj.startswith('+') and re.match(r'^\+\d+(?:[.,]\d+)?$', obj):
            return obj[1:]
        else:
            return obj
    else:
        return obj


def mark_merged_cables(excel_path, json_path='merged_list.json'):
    """
    Отмечает кабели из JSON-списка как "Объединён" в столбце 28 'Статус объединения'.
    Args:
        excel_path: путь к Excel-файлу базы данных
        json_path: путь к JSON-файлу со списком KKS (по умолчанию 'merged_list.json')
    
    Returns:
        int: количество отмеченных кабелей
    """
    # Проверяем, существует ли JSON-файл
    json_file = Path(json_path)
    if not json_file.exists():
        print(f"   ⚠️ Файл {json_path} не найден. Пропуск отметки объединённых кабелей.")
        return 0
    
    # Загружаем список KKS из JSON
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            merged_kks_list = json.load(f)
    except Exception as e:
        print(f"   ⚠️ Ошибка чтения {json_path}: {e}")
        return 0
    
    if not merged_kks_list:
        print(f"   ⚠️ Файл {json_path} пуст. Пропуск отметки объединённых кабелей.")
        return 0
    
    # Преобразуем список в множество для быстрого поиска
    merged_kks_set = set(str(kks).strip() for kks in merged_kks_list if kks)
    
    print(f"   📋 Загружено KKS для отметки: {len(merged_kks_set)}")
    
    # Загружаем Excel-файл
    wb = load_workbook(excel_path)
    sheet = wb.active
    
    # Находим индекс колонки 'ККС' и 'Статус объединения'
    kks_col = None
    status_col = None
    
    for col_idx, cell in enumerate(sheet[1], start=1):
        if cell.value and 'ККС' in str(cell.value):
            kks_col = col_idx
        if cell.value and 'Статус объединения' in str(cell.value):
            status_col = col_idx
    
    if kks_col is None:
        print(f"   ⚠️ Колонка 'ККС' не найдена. Пропуск отметки.")
        return 0
    
    if status_col is None:
        print(f"   ⚠️ Колонка 'Статус объединения' не найдена. Пропуск отметки.")
        return 0
    
    # Проходим по всем строкам и отмечаем совпадающие KKS
    marked_count = 0
    for row_idx in range(2, sheet.max_row + 1):
        kks_cell = sheet.cell(row=row_idx, column=kks_col)
        kks_value = str(kks_cell.value).strip() if kks_cell.value else ''
        
        if kks_value in merged_kks_set:
            status_cell = sheet.cell(row=row_idx, column=status_col)
            status_cell.value = 'Объединён'
            marked_count += 1
    
    # Сохраняем изменения
    wb.save(excel_path)
    print(f"   ✅ Отмечено кабелей как 'Объединён': {marked_count}")
    
    return marked_count


def check_start_end_trace(from_room, to_room, trace):
    """
    Проверяет, содержит ли трасса ККС кабельных конструкций для указанных зданий.
    Args:
        from_room: ККС помещения откуда (например, '07UBG13R013')
        to_room: ККС помещения куда (например, '07UBG13R009')
        trace: строка с описанием трассы
    Returns:
        bool: True если найдены соответствующие ККС кабельных конструкций, иначе False
    """

    # 1. Извлекаем ККС здания откуда и куда
    from_building_match = regular_KKS_building.search(str(from_room))
    to_building_match = regular_KKS_building.search(str(to_room))
    
    # Если не удалось извлечь ККС здания, возвращаем False
    if not from_building_match or not to_building_match:
        return False
    
    from_building = from_building_match.group()  # например, '07UBG'
    to_building = to_building_match.group()      # например, '07UBG'
    
    # 2. Находим все ККС кабельных конструкций в строке трассы
    if not trace:
        return False
    
    # Находим все совпадения с regular_cable_tray
    tray_matches = regular_cable_tray.findall(trace)
    
    if not tray_matches:
        return False
    
    # 3. Проверяем, содержат ли найденные ККС нужные здания
    found_from = False
    found_to = False
    
    for tray in tray_matches:
        if from_building in tray:
            found_from = True
        if to_building in tray:
            found_to = True
    
    # 4. Возвращаем результат в зависимости от совпадения зданий
    if from_building == to_building:    # Здания совпадают: достаточно одного совпадения
        return found_from or found_to
    else:                               # Здания разные: нужны оба
        return found_from and found_to


def build_cable_database(journals_dir, output_dir, progress_callback=None, source_type='СУПИР'):
    """
    Создаёт кабельную базу из журналов в указанной папке.
    Args:
        journals_dir: папка с журналами (.doc, .docx)
        output_dir: папка, в которую будет сохранена новая база
        progress_callback: функция для обновления прогресса (принимает current, total, message)
        source_type: источник данных ('СУПИР' или 'ВК')
    """
    print("\nЗапуск build_cable_database...")
    print(f"   📌 Источник журналов: {source_type}")

    # Загрузка границ зданий из JSON
    with open('KKS_building_bounds.json', 'r', encoding='utf-8') as f:
        building_bounds = json.load(f)

    # Поиск старой версии базы
    b_name = config.BASE_NAME
    current_version = current_version_finder(output_dir, b_name)
    
    # Получение списка всех журналов (.doc и .docx)
    journals = take_all_docx_from_dir(journals_dir)

    # journals_names = [j.stem for j in journals] if journals else []

    journals_names = []
    if journals:
        for j in journals:
            # очистка названия КЖ перед записью
            jrn = j.stem
            match = regular_journal_kks.search(jrn)
            if match:
                jrn = match.group()
            journals_names.append(jrn)
            jrn = ''

    total_journals = len(journals)
    
    # Отчёт о прогрессе: найдено журналов
    if progress_callback:
        progress_callback(0, total_journals, f"Найдено {total_journals} журналов")

    # Создание новой базы
    wb = Workbook()
    sheetWrite = wb.active
    sheetWrite.title = f"База данных вер. {current_version}"

    # Заголовки таблицы
    heads = config.HEADERS
    for head in heads:
        col = head[0] + 1
        sheetWrite.cell(row=1, column=col, value=f"{head[0]+1}. {head[1]}")
        sheetWrite.column_dimensions[get_column_letter(col)].width = head[2]
    sheetWrite.freeze_panes = 'A2'

    today = date.today()
    date_string = today.strftime("%Y.%m.%d")
    
    rowWrite = 2
    
    # ----- Копирование из старой версии базы -----
    if current_version > 1:
        old_path = Path(output_dir) / f"{b_name}{current_version - 1}.xlsx"
        if old_path.exists():
            rb = load_workbook(old_path, data_only=True)
            sheetRead = rb.active

            rows_to_copy = []
            for idx, row in enumerate(sheetRead.iter_rows(min_row=2, values_only=True), start=2):
                journal_name = str(row[0]) if row[0] else ''
                if journal_name and journal_name not in journals_names:
                    rows_to_copy.append(row)

            total_old = len(rows_to_copy)
            if progress_callback:
                progress_callback(0, total_journals + total_old, f"Перенос старых записей (всего {total_old})")
            
            for i, row_data in enumerate(rows_to_copy):
                for col_idx, value in enumerate(row_data, start=1):
                    sheetWrite.cell(row=rowWrite, column=col_idx, value=value)
                rowWrite += 1
                if progress_callback and i % 10 == 0:
                    progress_callback(i, total_journals + total_old, f"Перенос старых записей: {i}/{total_old}")

    # ----- Обработка новых журналов -----
    if journals:
        if progress_callback:
            progress_callback(0, total_journals, "Начало обработки журналов")
        
        for i, journal in enumerate(journals):
            try:
                if progress_callback:
                    progress_callback(i, total_journals, f"Обработка: {journal.stem}")
                
                # ========== КОНВЕРТАЦИЯ АВТОМАТИЧЕСКОЙ НУМЕРАЦИИ ==========
                # Сначала конвертируем нумерацию в текст
                from doc_utils import convert_numbering_to_text
                convert_numbering_to_text(str(journal))

                j_raw_content = extract_all_text_from_docx(Document(journal))
                if not j_raw_content:
                    if progress_callback:
                        progress_callback(i, total_journals, f"Нет данных в {journal.stem}")
                    continue
                
                # =================================
                # Получение ревизии
                revision = ''
                if 'revision_' in j_raw_content[-1]:
                    revision = j_raw_content[-1][-3:]
                    j_raw_content = j_raw_content[:-1]
                # =================================

                for raw_row in j_raw_content:
                    raw_row = remove_plus_from_numbers(raw_row)
                    
                    processed_row = None

                    if len(raw_row) > 6:
                        # очистка названия КЖ перед записью
                        jrn = journal.stem
                        match = regular_journal_kks.search(jrn)
                        if match:
                            jrn = match.group()
                        raw_row.insert(0, jrn)
                        jrn = ''

                        processed_row = row_parser(raw_row, building_bounds)

                    rows_processed = 0

                    if processed_row:
                        processed_row.append(date_string)
                        processed_row.append(current_version)
                        
                        for idx, val in enumerate(processed_row, start=1):
                            sheetWrite.cell(row=rowWrite, column=idx, value=val)
                        
                        sheetWrite.cell(row=rowWrite, column=24, value=revision)
                        sheetWrite.cell(row=rowWrite, column=27, value=source_type)
                        sheetWrite.cell(row=rowWrite, column=30, value=str(raw_row))
                        
                        # Расчёт минимальной длины кабеля
                        try:
                            # Функция для преобразования координаты (замена запятой на точку)
                            def to_float(val):
                                if not val:
                                    return None
                                return float(str(val).replace(',', '.'))
                            
                            length = processed_row[7]
                            if not length:
                                raise ValueError("Нет длины")
                            
                            length_val = float(length)
                            
                            from_x = to_float(processed_row[11])
                            from_y = to_float(processed_row[12])
                            from_z = to_float(processed_row[13])
                            to_x = to_float(processed_row[16])
                            to_y = to_float(processed_row[17])
                            to_z = to_float(processed_row[18])
                            
                            # Проверяем, что все координаты есть
                            if None in (from_x, from_y, from_z, to_x, to_y, to_z):
                                raise ValueError("Не все координаты")
                            
                            min_len = round(
                                abs(from_x - to_x) + 
                                abs(from_y - to_y) # + abs(from_z - to_z)
                            )
                            
                            sheetWrite.cell(row=rowWrite, column=31, value=min_len)
                            
                            # если фактическая длина меньше минимальной, то
                            if length_val < min_len:
                                sheetWrite.cell(row=rowWrite, column=32, value='Длина меньше минимальной')

                        except Exception:
                            # Если что-то пошло не так — просто пропускаем
                            pass
                        
                        # Проверка наличия кабельных конструкций здания начала и конца трассы.
                        # Если в трассировке нет конструкций из начала И конца, то это признак необходимости объединения
                        check_trace = processed_row[8]
                        check_from_room = processed_row[9]
                        check_to_room = processed_row[14]
                        
                        if check_trace and check_from_room and check_to_room:
                            # Извлекаем ККС зданий
                            from_match = regular_KKS_building.search(str(check_from_room))
                            to_match = regular_KKS_building.search(str(check_to_room))
                            
                            if from_match and to_match:
                                from_building = from_match.group()
                                to_building = to_match.group()
                                
                                # Если здания одинаковые
                                if from_building == to_building:
                                    # Проверяем, есть ли в трассе ККС кабельных конструкций
                                    tray_matches = regular_cable_tray.findall(check_trace)
                                    if not tray_matches:
                                        # Нет ни одной кабельной конструкции — пропускаем проверку, это прокладка по месту
                                        pass
                                    # else:
                                    #     # Есть кабельные конструкции, но не для этого здания?
                                    #     if not check_start_end_trace(check_from_room, check_to_room, check_trace):
                                    #         current = sheetWrite.cell(row=rowWrite, column=32).value or ''
                                    #         if current:
                                    #             sheetWrite.cell(row=rowWrite, column=32, value=current + '; Да, трасса не полная')
                                    #         else:
                                    #             sheetWrite.cell(row=rowWrite, column=32, value='Да, трасса не полная')
                                else:
                                    # Здания разные
                                    if not check_start_end_trace(check_from_room, check_to_room, check_trace):
                                        current = sheetWrite.cell(row=rowWrite, column=32).value or ''
                                        if current:
                                            sheetWrite.cell(row=rowWrite, column=32, value=current + '; трасса не полная')
                                        else:
                                            sheetWrite.cell(row=rowWrite, column=32, value='трасса не полная')

                        # Проверка наличия ссылки на кабельный журнал в трассе
                        # Если есть, то в графу 32, 'Ответная часть (из КЖ)' внести ККС журнала, в 31, 'Требования к объединению' - еще комментарий
                        journal_matches = regular_journal_kks_short.findall(str(check_trace))
                        if journal_matches:
                            # Удаляем дубликаты через set и объединяем через пробел
                            j_in_trace = ' '.join(sorted(set(journal_matches), key=journal_matches.index))
                            
                            sheetWrite.cell(row=rowWrite, column=33, value=j_in_trace)
                            
                            current = sheetWrite.cell(row=rowWrite, column=32).value or ''
                            separator = '; ' if current else ''
                            sheetWrite.cell(row=rowWrite, column=32, value=f"{current}{separator}есть ответная часть")
                            
                        rowWrite += 1
                        rows_processed += 1

                if progress_callback:
                    progress_callback(i + 1, total_journals, f"Готово: {journal.stem} ({rows_processed} строк)")

            except Exception as e:
                print(f"Ошибка обработки {journal.stem}: {e}")
                if progress_callback:
                    progress_callback(i + 1, total_journals, f"Ошибка в {journal.stem}: {str(e)[:50]}", is_error=True)
    


    # ========== ПРОВЕРКА ЖУРНАЛОВ В ОТВЕТНОЙ ЧАСТИ ==========
    # 1. Составляем список set всех ККС журналов во всей таблице (первый столбец)
    all_journals_in_table = set()
    for row in range(2, rowWrite):  # rowWrite - это следующая строка после последней заполненной
        all_journals_in_table.add(sheetWrite.cell(row=row, column=1).value)
    print(f"   📋 Найдено уникальных журналов в таблице: {len(all_journals_in_table)}")
    
    # 2. Проверяем кабели, у которых есть ККС журнала в столбце 33
    for row in range(2, rowWrite):
        ref_jrn = sheetWrite.cell(row=row, column=33).value
        if ref_jrn:
            ref_journals = str(ref_jrn).split()
            missing = []
            
            for r in ref_journals:
                # Точное совпадение или подстрока
                if r in all_journals_in_table:
                    continue
                # Проверяем, содержится ли r как подстрока в каком-либо журнале
                if any(r in journal for journal in all_journals_in_table if journal):
                    continue
                missing.append(r)
            
            if missing:
                sheetWrite.cell(row=row, column=34, value='нет в базе')
            else:
                sheetWrite.cell(row=row, column=34, value='есть в базе')
    
    # ========== КОНЕЦ БЛОКА ПРОВЕРКИ ==========


    # Применяем автофильтр и сохраняем
    sheetWrite.auto_filter.ref = sheetWrite.dimensions
    output_path = Path(output_dir) / f"{b_name}{current_version}.xlsx"
    wb.save(output_path)
    
    # ========== ОТМЕТКА ОБЪЕДИНЁННЫХ КАБЕЛЕЙ ==========
    # Пытаемся отметить кабели из merged_list.json как "Объединён"
    print("\n🏷️ Отметка объединённых кабелей...")
    mark_merged_cables(output_path, 'merged_list.json')
    # ========== КОНЕЦ БЛОКА ==========
    
    if progress_callback:
        progress_callback(total_journals, total_journals, "Сохранение базы данных")
    
    print(f"\n✅ База данных сохранена: {output_path}")
    print("🏁 Готово!")