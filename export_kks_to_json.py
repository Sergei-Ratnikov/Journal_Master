import json
from openpyxl import load_workbook
import re
import config
from pathlib import Path

def cleanCyrFromLat(lineIn):
    """Замена русских букв на латинские"""
    return str(lineIn).replace('А', 'A').replace('В', 'B').replace('С', 'C').replace('Е', 'E').replace('Н', 'H').replace('К', 'K').replace('М', 'M').replace('О', 'O').replace('Р', 'P').replace('Т', 'T').replace('Х', 'X').replace(' ', '').replace(',', '.')

def parse_route_to_bounds(route_string):
    """
    Извлекает из строки Route координаты квадрата и возвращает расширенные границы
    """
    pattern = re.compile(r'(\d{1,2})\s*([EN])')
    matches = pattern.findall(route_string.upper())
    
    if len(matches) < 2:
        return {
            'is_valid': False,
            'original': route_string,
            'x_min': None,
            'x_max': None,
            'y_min': None,
            'y_max': None
        }
    
    e_val = None
    n_val = None
    
    for num, axis in matches:
        value = int(num) * 100
        if axis == 'E':
            e_val = value
        elif axis == 'N':
            n_val = value
    
    if e_val is None or n_val is None:
        return {
            'is_valid': False,
            'original': route_string,
            'x_min': None,
            'x_max': None,
            'y_min': None,
            'y_max': None
        }
    
    # Расширяем диапазон на ±100 метров
    return {
        'is_valid': True,
        'original': route_string,
        'x_min': n_val,
        'x_max': n_val + 100,
        'y_min': e_val,
        'y_max': e_val + 100
    }


def export_kks_to_json(excel_path, json_path='building_bounds.json'):
    """
    Читает KKS.xlsx и сохраняет building_bounds в JSON
    """
    if not Path(excel_path).exists():
        print(f"❌ Файл не найден: {excel_path}")
        return
    
    building_bounds = {}
    
    wb = load_workbook(excel_path, data_only=True)
    sheet = wb.active
    
    row_count = 0
    saved_count = 0
    
    for row in range(2, sheet.max_row + 1):
        kks = str(sheet.cell(row=row, column=1).value).strip()
        route = str(sheet.cell(row=row, column=4).value).strip()
        row_count += 1
        
        if not kks or not config.regular_KKS_building.search(kks):
            continue
        
        kks_clean = cleanCyrFromLat(kks)
        bounds = parse_route_to_bounds(route)
        building_bounds[kks_clean] = bounds
        saved_count += 1
    
    # Сохраняем в JSON
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(building_bounds, f, ensure_ascii=False, indent=2)
    
    print(f"\n📊 Статистика:")
    print(f"   Всего строк в Excel: {row_count}")
    print(f"   Сохранено зданий: {saved_count}")
    print(f"   Пропущено (без KKS здания): {row_count - saved_count}")
    print(f"\n✅ Файл сохранён: {json_path}")


# if __name__ == "__main__":
#     # Укажите путь к вашему KKS.xlsx


export_kks_to_json('JournalMaster\KKS.xlsx')