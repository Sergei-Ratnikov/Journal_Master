# read_xlsx.py
"""
Все функции для работы с Excel-журналами (.xlsx)
Извлечение таблиц, конвертация, перемещение файлов
Формат выхода такой же, как у read_docx.py
"""

import os
import shutil
import re
from pathlib import Path
from openpyxl import load_workbook
import config


# ========== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ==========

def all_deep_empty(lst):
    """Рекурсивно проверяет, пусты ли все вложенные списки."""
    if not isinstance(lst, list):
        return False
    if not lst:
        return True
    return all(all_deep_empty(item) for item in lst)


# ========== ФУНКЦИИ ДЛЯ КОНВЕРТАЦИИ И ПЕРЕМЕЩЕНИЯ ==========

def move_file(source_path, destination_dir):
    """
    Переносит файл из source_path в папку destination_dir
    """
    if not os.path.exists(source_path):
        print(f"Файл не найден: {source_path}")
        return False

    try:    
        os.makedirs(destination_dir, exist_ok=True)
        filename = os.path.basename(source_path)
        destination_path = os.path.join(destination_dir, filename)
        shutil.move(source_path, destination_path)
        print(f"Файл перемещён: {source_path} → {destination_path}")
        return True
    except Exception as e:
        print(f"Ошибка при переносе: {e}")
        return False


def take_all_xlsx_from_dir(journals_directory):
    """
    Получение списка всех .xlsx файлов из указанной директории.
    Возвращает: (список_файлов, список_перемещённых)
    """
    source_dir = Path(journals_directory).resolve()
    
    xlsx_files = list(source_dir.glob('*.xlsx'))
    
    # Исключаем файлы базы данных (Cable base ver.*.xlsx)
    xlsx_files = [f for f in xlsx_files if not f.name.startswith('Cable base ver.')]
    
    return xlsx_files, []  # вторым возвращаем пустой список (для совместимости с read_docx)


def extract_cell_text(cell):
    """Извлекает текст из ячейки Excel."""
    if cell is None:
        return ''
    if cell.value is None:
        return ''
    if isinstance(cell.value, (int, float)):
        return str(cell.value).replace('.', ',').strip()
    return str(cell.value).strip()


def find_data_start_row(sheet):
    """
    Находит первую строку с данными (где есть номер кабеля).
    Ищем в столбце A (индекс 1) значение, похожее на номер кабеля (число или число.число)
    """
    import re
    for row_idx in range(1, min(sheet.max_row + 1, 100)):
        cell = sheet.cell(row=row_idx, column=1)
        val = extract_cell_text(cell)
        if val and re.match(r'^\d+(?:[.,]\d+)?$', val):
            return row_idx
    return None


def extract_all_text_from_xlsx(xlsx_path):
    """
    Извлекает данные из Excel-журнала.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    sheet = wb.active
    
    start_row = find_data_start_row(sheet)
    if start_row is None:
        print(f"  ⚠️ Не найдены данные в {xlsx_path}")
        return []
    
    result = []
    current_cable = None
    current_rows = []
    
    for row_idx in range(start_row, sheet.max_row + 1):
        code_cell = sheet.cell(row=row_idx, column=1)
        code = extract_cell_text(code_cell)
        
        if code and re.match(r'^\d+(?:[.,]\d+)?$', code):
            if current_cable is not None and current_rows:
                result.append(process_cable_rows(current_cable, current_rows, sheet))  # ← добавить sheet
            
            current_cable = code
            current_rows = [row_idx]
        else:
            if current_cable is not None:
                current_rows.append(row_idx)
    
    if current_cable is not None and current_rows:
        result.append(process_cable_rows(current_cable, current_rows, sheet))  # ← добавить sheet
    
    return result


def process_cable_rows(cable_number, row_indices, sheet):
    """
    Обрабатывает все строки, относящиеся к одному кабелю.
    Собирает данные из всех строк диапазона.
    """
    # Собираем данные из всех колонок по строкам
    col_data = {
        'B': [],   # ККС, марка, сечение
        'C': [],   # группа, класс безопасности
        'FROM': [],  # откуда (D, E, F)
        'TO': [],    # куда (G, H, I)
        'J': [],   # длина
        'K': [],   # трасса (склеивается)
    }
    
    for row_idx in row_indices:
        # B: ККС, марка, сечение
        val = extract_cell_text(sheet.cell(row=row_idx, column=2))
        if val:
            # Разбиваем по \n и добавляем все части
            col_data['B'].extend([v.strip() for v in val.split('\n') if v.strip()])
        
        # C: группа, класс безопасности (каждая ячейка — отдельный элемент)
        val = extract_cell_text(sheet.cell(row=row_idx, column=3))
        if val:
            # Разбиваем по \n, каждый фрагмент — отдельный элемент
            for v in val.split('\n'):
                v = v.strip()
                if v:
                    col_data['C'].append(v)
        

        # D, E, F: откуда (помещение + оборудование + координаты)
        # Просто собираем все непустые строки из D, E, F
        for col in [4, 5, 6]:
            val = extract_cell_text(sheet.cell(row=row_idx, column=col))
            if val:
                for v in val.split('\n'):
                    v = v.strip()
                    if v:
                        col_data['FROM'].append(v)
        
        # G, H, I: куда (помещение + оборудование + координаты)
        for col in [7, 8, 9]:
            val = extract_cell_text(sheet.cell(row=row_idx, column=col))
            if val:
                for v in val.split('\n'):
                    v = v.strip()
                    if v:
                        col_data['TO'].append(v)

        # J: длина
        val = extract_cell_text(sheet.cell(row=row_idx, column=10))
        if val:
            col_data['J'].extend([v.strip() for v in val.split('\n') if v.strip()])
        
        # K: трасса (склеиваем всё в одну строку)
        val = extract_cell_text(sheet.cell(row=row_idx, column=11))
        if val:
            # Заменяем \n на пробел и добавляем
            col_data['K'].append(val.replace('\n', ' ').strip())
        # Формируем строку в формате, совместимом с read_docx.py
    
    # 0. Журнал (будет добавлен позже в excel_utils)
    # 1. Номер кабеля
    cell_1 = [cable_number]
    
    # 2. ККС кабеля + марка + сечение (всё из колонки B в виде списка)
    cell_2 = col_data['B'] if col_data['B'] else ['']
    
    # 3. Группа + класс безопасности (каждый элемент отдельно)
    cell_3 = col_data['C'] if col_data['C'] else ['']
    
    # 4. Блок "Откуда" (все строки из D, E, F)
    cell_4 = col_data['FROM'] if col_data['FROM'] else ['']
    
    # 5. Блок "Куда" (все строки из G, H, I)
    cell_5 = col_data['TO'] if col_data['TO'] else ['']
    
    # 6. Длина
    length = ' '.join(col_data['J']) if col_data['J'] else ''
    cell_6 = [length] if length else ['']
    
    # 7. Трасса (склеиваем все части через пробел)
    trace = ' '.join(col_data['K']) if col_data['K'] else ''
    cell_7 = [trace] if trace else ['']
    
    # Собираем итоговую строку
    cable_row = [cell_1, cell_2, cell_3, cell_4, cell_5, cell_6, cell_7]
    
    return cable_row