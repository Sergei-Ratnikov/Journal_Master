"""
Refresh_answer_cable.py
Отдельная утилита для проверки наличия кабелей в ответных журналах.

Логика работы:
    1. Открывает Cable base ver.*.xlsx
    2. Строит словарь {ККС_журнала: [ККС_кабеля1, ККС_кабеля2, ...]} из всех строк
    3. Находит строки с записями в столбце AH (Ответная часть из КЖ)
    4. Извлекает ККС журнала из строки по regular_journal_kks_short
    5. Проверяет наличие ККС кабеля в найденном журнале
    6. Записывает "есть" или "нет" в столбец AI
    7. Сохраняет и закрывает файл
"""

import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import threading
import sys
import os

# Добавляем путь к проекту для импорта config
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import regular_journal_kks_short


class CableCheckerGUI:
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Проверка наличия кабелей в ответных журналах")
        self.root.geometry("600x500")
        self.root.resizable(False, False)
        self.root.configure(bg="#f0f0f0")
        
        self.excel_path = tk.StringVar()
        self.is_running = False
        
        self.setup_ui()
        self.center_window()
    
    def center_window(self):
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def setup_ui(self):
        # Заголовок
        title_frame = tk.Frame(self.root, bg="#2c3e50", height=80)
        title_frame.pack(fill="x")
        title_frame.pack_propagate(False)
        
        tk.Label(title_frame, text="ПРОВЕРКА КАБЕЛЕЙ В ОТВЕТНЫХ ЖУРНАЛАХ", 
                font=("Arial", 14, "bold"), fg="white", bg="#2c3e50").pack(expand=True)
        tk.Label(title_frame, text="Поиск кабелей по ККС в указанных журналах",
                font=("Arial", 10), fg="#bdc3c7", bg="#2c3e50").pack()
        
        # Основная область
        main_frame = tk.Frame(self.root, padx=20, pady=20, bg="#f0f0f0")
        main_frame.pack(fill="both", expand=True)
        
        # Выбор файла
        file_frame = tk.LabelFrame(main_frame, text="📄 Файл базы данных", 
                                   padx=10, pady=10, font=("Arial", 10, "bold"), bg="#f0f0f0")
        file_frame.pack(fill="x", pady=(0, 15))
        
        tk.Entry(file_frame, textvariable=self.excel_path, 
                font=("Arial", 10), state="readonly", bg="white").pack(side="left", fill="x", expand=True, padx=(0, 10))
        tk.Button(file_frame, text="Выбрать файл", command=self.select_file,
                 bg="#3498db", fg="white", font=("Arial", 9), padx=10, cursor="hand2").pack(side="right")
        
        # Информация
        info_frame = tk.LabelFrame(main_frame, text="📊 Информация", 
                                   padx=10, pady=10, font=("Arial", 10, "bold"), bg="#f0f0f0")
        info_frame.pack(fill="x", pady=(0, 15))
        
        self.info_label = tk.Label(info_frame, text="Выберите файл базы данных",
                                   font=("Arial", 9), bg="#f0f0f0", fg="#7f8c8d")
        self.info_label.pack()
        
        # Прогресс
        progress_frame = tk.LabelFrame(main_frame, text="📊 Прогресс", 
                                       padx=10, pady=10, font=("Arial", 10, "bold"), bg="#f0f0f0")
        progress_frame.pack(fill="x", pady=(0, 15))
        
        self.progress_bar = ttk.Progressbar(progress_frame, mode="determinate")
        self.progress_bar.pack(fill="x", pady=(0, 5))
        
        self.progress_label = tk.Label(progress_frame, text="Ожидание запуска...",
                                       font=("Arial", 9), bg="#f0f0f0", fg="#7f8c8d")
        self.progress_label.pack()
        
        # Кнопка запуска
        self.run_btn = tk.Button(main_frame, text="🚀 ЗАПУСТИТЬ ПРОВЕРКУ",
                                 command=self.run_check, bg="#27ae60", fg="white",
                                 font=("Arial", 12, "bold"), height=2, cursor="hand2")
        self.run_btn.pack(fill="x", pady=(0, 15))
        
        # Статус
        self.status_label = tk.Label(main_frame, text="Готов к работе.",
                                     font=("Arial", 9), fg="#7f8c8d", bg="#f0f0f0")
        self.status_label.pack()
    
    def select_file(self):
        file_path = filedialog.askopenfilename(
            title="Выберите файл базы данных",
            filetypes=[("Excel files", "Cable base ver.*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.excel_path.set(file_path)
            self.info_label.config(text=f"Файл: {os.path.basename(file_path)}")
            self.update_status(f"Выбран файл: {os.path.basename(file_path)}")
    
    def update_status(self, message, is_error=False):
        if is_error:
            self.status_label.config(text=f"❌ {message}", fg="#e74c3c")
        else:
            self.status_label.config(text=message, fg="#7f8c8d")
        self.root.update_idletasks()
    
    def update_progress(self, current, total, message):
        if total > 0:
            percent = int((current / total) * 100)
            self.progress_bar["value"] = percent
            self.progress_label.config(text=f"{percent}% ({current}/{total}) - {message}")
        self.root.update_idletasks()
    
    def run_check(self):
        if self.is_running:
            return
        
        excel_path = self.excel_path.get()
        if not excel_path:
            messagebox.showerror("Ошибка", "Выберите файл базы данных!")
            return
        
        if not Path(excel_path).exists():
            messagebox.showerror("Ошибка", f"Файл не найден:\n{excel_path}")
            return
        
        self.is_running = True
        self.run_btn.config(state="disabled", bg="#95a5a6")
        self.progress_bar["value"] = 0
        self.progress_label.config(text="Подготовка...")
        self.update_status("⏳ Выполняется проверка...")
        
        thread = threading.Thread(target=self._process, daemon=True)
        thread.start()
    
    def _process(self):
        try:
            excel_path = self.excel_path.get()
            
            self.update_progress(0, 100, "Загрузка файла...")
            wb = load_workbook(excel_path, data_only=True)
            sheet = wb.active
            
            # Определяем колонки
            col_journal = 1    # A
            col_kks = 3        
            col_response = 34  # AG (Ответная часть из КЖ)
            col_result = 35    # AI (Результат)
            
            # ========== 1. СТРОИМ СЛОВАРЬ {ККС_журнала: [ККС_кабеля1, ...]} ==========
            self.update_progress(5, 100, "Построение индекса кабелей...")
            journal_cable_map = {}
            max_row = sheet.max_row
            
            for row_idx in range(2, max_row + 1):
                journal = sheet.cell(row=row_idx, column=col_journal).value
                kks = sheet.cell(row=row_idx, column=col_kks).value
                if journal and kks:
                    journal_str = str(journal).strip()
                    kks_str = str(kks).strip()
                    if journal_str not in journal_cable_map:
                        journal_cable_map[journal_str] = []
                    journal_cable_map[journal_str].append(kks_str)
                
                if row_idx % 1000 == 0:
                    progress = 5 + int((row_idx / max_row) * 30)
                    self.update_progress(progress, 100, f"Индекс: {row_idx}/{max_row} строк")
            
            print(f"   📋 Построен индекс: {len(journal_cable_map)} журналов")
            self.update_progress(40, 100, f"Индекс построен: {len(journal_cable_map)} журналов")
            
            # ========== 2. ИЩЕМ СТРОКИ С ОТВЕТНОЙ ЧАСТЬЮ ==========
            self.update_progress(45, 100, "Поиск строк с ответной частью...")
            response_rows = []
            for row_idx in range(2, max_row + 1):
                response_val = sheet.cell(row=row_idx, column=col_response).value
                if response_val:
                    response_rows.append(row_idx)
            
            print(f"   📋 Найдено строк с ответной частью: {len(response_rows)}")
            self.update_progress(50, 100, f"Найдено {len(response_rows)} строк с ответной частью")
            
            # ========== 3. ПРОВЕРЯЕМ КАЖДУЮ СТРОКУ ==========
            total = len(response_rows)
            found_count = 0
            not_found_count = 0
            
            for i, row_idx in enumerate(response_rows):
                # Получаем ККС кабеля из столбца C (индекс 2 в openpyxl)
                cable_kks = sheet.cell(row=row_idx, column=col_kks).value
                cable_kks = str(cable_kks).strip() if cable_kks else ''
                
                # Получаем текст ответной части
                response_text = sheet.cell(row=row_idx, column=col_response).value
                response_text = str(response_text).strip() if response_text else ''
                
                # Извлекаем ККС журнала по регулярному выражению
                journal_kks_found = None
                if response_text:
                    match = regular_journal_kks_short.search(response_text)
                    if match:
                        journal_kks_found = match.group()
                
                # Ищем кабель в словаре
                found = False
                if journal_kks_found and cable_kks:
                    # Проверяем все возможные варианты (точное совпадение или префикс)
                    for full_journal in journal_cable_map.keys():
                        if full_journal == journal_kks_found or full_journal.startswith(journal_kks_found):
                            if cable_kks in journal_cable_map[full_journal]:
                                found = True
                                break
                
                # Записываем результат
                if found:
                    sheet.cell(row=row_idx, column=col_result, value='есть')
                    found_count += 1
                else:
                    sheet.cell(row=row_idx, column=col_result, value='нет')
                    not_found_count += 1
                
                progress = 50 + int(((i + 1) / total) * 45)
                self.update_progress(progress, 100, f"Строка {i+1}/{total}: {cable_kks} → {'есть' if found else 'нет'}")
            
            # ========== 4. СОХРАНЯЕМ ==========
            self.update_progress(98, 100, "Сохранение файла...")
            wb.save(excel_path)
            
            self.root.after(0, self._on_success, found_count, not_found_count, len(response_rows))
            
        except Exception as e:
            self.root.after(0, self._on_error, str(e))
    
    def _on_success(self, found_count, not_found_count, total_rows):
        self.is_running = False
        self.run_btn.config(state="normal", bg="#27ae60")
        self.progress_bar["value"] = 100
        self.progress_label.config(text=f"Готово! Найдено: {found_count}, Не найдено: {not_found_count}")
        self.update_status(f"✅ Проверка завершена! Найдено: {found_count}, Не найдено: {not_found_count}")
        
        messagebox.showinfo("Готово!", 
            f"✅ Проверка завершена!\n\n"
            f"📄 Файл: {os.path.basename(self.excel_path.get())}\n"
            f"📊 Всего строк с ответной частью: {total_rows}\n"
            f"✅ Найдено кабелей: {found_count}\n"
            f"❌ Не найдено: {not_found_count}")
    
    def _on_error(self, error_msg):
        self.is_running = False
        self.run_btn.config(state="normal", bg="#27ae60")
        self.progress_label.config(text=f"Ошибка: {error_msg[:50]}")
        self.update_status(f"Ошибка: {error_msg}", is_error=True)
        messagebox.showerror("Ошибка", f"Произошла ошибка:\n\n{error_msg}")
    
    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = CableCheckerGUI()
    app.run()