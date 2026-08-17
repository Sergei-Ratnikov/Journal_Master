"""
file_checker.py
Утилита для проверки наличия файлов в папке по списку из Excel.
"""

import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
from openpyxl import load_workbook
import threading


class FileCheckerGUI:
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Проверка наличия файлов")
        self.root.geometry("600x480")
        self.root.resizable(False, False)
        self.root.configure(bg="#f0f0f0")
        
        self.excel_path = tk.StringVar()
        self.folder_path = tk.StringVar()
        self.is_running = False
        
        self.setup_ui()
    
    def setup_ui(self):
        # Заголовок
        title_frame = tk.Frame(self.root, bg="#2c3e50", height=70)
        title_frame.pack(fill="x")
        title_frame.pack_propagate(False)
        
        tk.Label(title_frame, text="ПРОВЕРКА НАЛИЧИЯ ФАЙЛОВ", 
                font=("Arial", 16, "bold"), fg="white", bg="#2c3e50").pack(expand=True)
        tk.Label(title_frame, text="Поиск файлов в папке по списку из Excel",
                font=("Arial", 10), fg="#bdc3c7", bg="#2c3e50").pack()
        
        # Основная область
        main_frame = tk.Frame(self.root, padx=20, pady=20, bg="#f0f0f0")
        main_frame.pack(fill="both", expand=True)
        
        # Excel файл
        excel_frame = tk.LabelFrame(main_frame, text="📄 Excel файл", 
                                    padx=10, pady=10, font=("Arial", 10, "bold"), bg="#f0f0f0")
        excel_frame.pack(fill="x", pady=(0, 15))
        
        tk.Entry(excel_frame, textvariable=self.excel_path, 
                font=("Arial", 10), state="readonly", bg="white").pack(side="left", fill="x", expand=True, padx=(0, 10))
        tk.Button(excel_frame, text="Выбрать файл", command=self.select_excel,
                 bg="#3498db", fg="white", font=("Arial", 9), padx=10, cursor="hand2").pack(side="right")
        
        # Папка
        folder_frame = tk.LabelFrame(main_frame, text="📁 Папка для проверки", 
                                     padx=10, pady=10, font=("Arial", 10, "bold"), bg="#f0f0f0")
        folder_frame.pack(fill="x", pady=(0, 15))
        
        tk.Entry(folder_frame, textvariable=self.folder_path, 
                font=("Arial", 10), state="readonly", bg="white").pack(side="left", fill="x", expand=True, padx=(0, 10))
        tk.Button(folder_frame, text="Выбрать папку", command=self.select_folder,
                 bg="#3498db", fg="white", font=("Arial", 9), padx=10, cursor="hand2").pack(side="right")
        
        # Информация
        info_frame = tk.LabelFrame(main_frame, text="📊 Информация", 
                                   padx=10, pady=10, font=("Arial", 10, "bold"), bg="#f0f0f0")
        info_frame.pack(fill="x", pady=(0, 15))
        
        self.info_label = tk.Label(info_frame, text="Выберите Excel-файл и папку для проверки",
                                   font=("Arial", 9), bg="#f0f0f0", fg="#7f8c8d")
        self.info_label.pack()
        
        # Прогресс
        progress_frame = tk.LabelFrame(main_frame, text="📊 Прогресс", 
                                       padx=10, pady=10, font=("Arial", 10, "bold"), bg="#f0f0f0")
        progress_frame.pack(fill="x", pady=(0, 15))
        
        self.progress_bar = ttk.Progressbar(progress_frame, mode="determinate")
        self.progress_bar.pack(fill="x", pady=(0, 5))
        
        self.progress_label = tk.Label(progress_frame, text="Ожидание запуска...",
                                       font=("Arial", 9), bg="#f0f0f0", fg="#7f8c8d")
        self.progress_label.pack()
        
        # ========== КНОПКА ЗАПУСКА ==========
        self.run_btn = tk.Button(
            main_frame, 
            text="🚀 ЗАПУСТИТЬ ПРОВЕРКУ",
            command=self.run_check, 
            bg="#27ae60", 
            fg="white",
            font=("Arial", 12, "bold"), 
            height=2, 
            cursor="hand2"
        )
        self.run_btn.pack(fill="x", pady=(0, 15))
        # ====================================
        
        # Статус
        self.status_label = tk.Label(main_frame, text="Готов к работе.",
                                     font=("Arial", 9), fg="#7f8c8d", bg="#f0f0f0")
        self.status_label.pack()
    
    def select_excel(self):
        file_path = filedialog.askopenfilename(
            title="Выберите Excel-файл",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.excel_path.set(file_path)
            self.update_info()
    
    def select_folder(self):
        folder_path = filedialog.askdirectory(title="Выберите папку для проверки")
        if folder_path:
            self.folder_path.set(folder_path)
            self.update_info()
    
    def update_info(self):
        excel = self.excel_path.get()
        folder = self.folder_path.get()
        if excel and folder:
            self.info_label.config(text=f"Файл: {os.path.basename(excel)} | Папка: {folder}")
    
    def update_status(self, message, is_error=False):
        if is_error:
            self.status_label.config(text=f"❌ {message}", fg="#e74c3c")
        else:
            self.status_label.config(text=message, fg="#7f8c8d")
        self.root.update_idletasks()
    
    def update_progress(self, current, total, message):
        if total > 0:
            percent = int((current / total) * 100)
            self.progress_bar["value"] = percent
            self.progress_label.config(text=f"{percent}% ({current}/{total}) - {message}")
        self.root.update_idletasks()
    
    def run_check(self):
        if self.is_running:
            return
        
        excel_path = self.excel_path.get()
        folder_path = self.folder_path.get()
        
        if not excel_path:
            messagebox.showerror("Ошибка", "Выберите Excel-файл!")
            return
        
        if not folder_path:
            messagebox.showerror("Ошибка", "Выберите папку для проверки!")
            return
        
        if not os.path.exists(excel_path):
            messagebox.showerror("Ошибка", f"Файл не найден: {excel_path}")
            return
        
        if not os.path.exists(folder_path):
            messagebox.showerror("Ошибка", f"Папка не найдена: {folder_path}")
            return
        
        self.is_running = True
        self.run_btn.config(state="disabled", bg="#95a5a6")
        self.progress_bar["value"] = 0
        self.progress_label.config(text="Подготовка...")
        self.update_status("⏳ Выполняется проверка...")
        
        thread = threading.Thread(target=self._process, daemon=True)
        thread.start()
    
    def _process(self):
        try:
            excel_path = self.excel_path.get()
            folder_path = self.folder_path.get()
            
            self.update_progress(0, 100, "Загрузка Excel...")
            wb = load_workbook(excel_path)
            sheet = wb.active
            
            self.update_progress(10, 100, "Сканирование папки...")
            files = [f.name for f in Path(folder_path).iterdir() if f.is_file()]
            
            max_col = sheet.max_column
            if max_col < 2:
                sheet.cell(row=1, column=2, value="Наличие")
            
            total_rows = sheet.max_row
            found_count = 0
            
            for row_idx in range(2, total_rows + 1):
                cell = sheet.cell(row=row_idx, column=1)
                search_text = str(cell.value).strip() if cell.value else ''
                
                if not search_text:
                    sheet.cell(row=row_idx, column=2, value="")
                    continue
                
                found = False
                for f in files:
                    if search_text in f:
                        found = True
                        break
                
                if found:
                    sheet.cell(row=row_idx, column=2, value="есть")
                    found_count += 1
                else:
                    sheet.cell(row=row_idx, column=2, value="нет")
                
                progress = 20 + int((row_idx - 1) / total_rows * 70)
                self.update_progress(progress, 100, f"Строка {row_idx-1}/{total_rows-1}")
            
            self.update_progress(95, 100, "Сохранение...")
            wb.save(excel_path)
            
            self.root.after(0, self._on_success, found_count, total_rows - 1)
            
        except Exception as e:
            self.root.after(0, self._on_error, str(e))
    
    def _on_success(self, found_count, total_rows):
        self.is_running = False
        self.run_btn.config(state="normal", bg="#27ae60")
        self.progress_bar["value"] = 100
        self.progress_label.config(text=f"Готово! Найдено: {found_count} из {total_rows}")
        self.update_status(f"✅ Проверка завершена! Найдено {found_count} файлов из {total_rows}")
        
        messagebox.showinfo("Готово!", 
            f"✅ Проверка завершена!\n\n"
            f"📄 Файл: {os.path.basename(self.excel_path.get())}\n"
            f"📁 Папка: {self.folder_path.get()}\n"
            f"📊 Найдено: {found_count} из {total_rows}\n"
            f"💾 Файл сохранён")
    
    def _on_error(self, error_msg):
        self.is_running = False
        self.run_btn.config(state="normal", bg="#27ae60")
        self.progress_label.config(text=f"Ошибка: {error_msg[:50]}")
        self.update_status(f"Ошибка: {error_msg}", is_error=True)
        messagebox.showerror("Ошибка", f"Произошла ошибка:\n\n{error_msg}")
    
    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = FileCheckerGUI()
    app.run()