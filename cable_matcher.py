# cable_matcher.py
"""
Поиск ответных частей кабелей между журналами.

Логика работы:
    1. Загружается Excel-база, созданная excel_utils.py (файл Cable base ver.*.xlsx).
    2. Пользователь вводит ККС журнала, требующего объединения.
    3. Просматриваются все кабели из этого журнала.
    4. Для каждого кабеля проверяется:
       - Статус объединения (колонка 28) пуст
       - Требования к объединению (колонка 32) не пусты
    5. Если условия выполнены → ищем во всей базе ВСЕ кабели с тем же ККС (кроме самого себя)
    6. Для каждого найденного ответного кабеля проверяются проблемы:
       - Марка не совпадает
       - Помещения "Откуда"/"Куда" не совпадают
       - Оборудование "Откуда"/"Куда" не совпадает
       - Трассы вложены друг в друга
    7. Результат записывается в новый Excel-файл.
"""

import sys
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

# ========== КОНФИГУРАЦИЯ ==========
HEADERS = [
    [0, '№ п/п', 8],
    [1, 'ККС кабеля', 25],
    [2, 'Исходный журнал (+ источник, дата)', 50],
    [3, 'Ответный журнал (+ источник, дата)', 50],
    [4, 'Примечание', 40],
    [5, 'Проблемы', 40]
]


def get_column_index(sheet, header_name):
    """Находит индекс колонки по её названию."""
    for col_idx, cell in enumerate(sheet[1], start=1):
        if cell.value and header_name in str(cell.value):
            return col_idx
    return None


def get_journal_info(sheet, journal_kks, row_idx=None):
    """
    Получает информацию о журнале: источник и дату.
    Если передан row_idx, берёт данные из этой строки.
    Иначе ищет первую строку с данным журналом.
    """
    source_col = get_column_index(sheet, 'Источник информации')
    date_col = get_column_index(sheet, 'Дата ревизии')
    
    if source_col is None or date_col is None:
        return '-', '-'
    
    if row_idx is not None:
        source = sheet.cell(row=row_idx, column=source_col).value or '-'
        date = sheet.cell(row=row_idx, column=date_col).value or '-'
        return str(source).strip(), str(date).strip()
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row and row[0] and str(row[0]).strip() == journal_kks:
            source = row[source_col - 1] if source_col <= len(row) else '-'
            date = row[date_col - 1] if date_col <= len(row) else '-'
            return str(source).strip() if source else '-', str(date).strip() if date else '-'
    
    return '-', '-'


def find_all_matching_cables(sheet, cable_kks, source_journal):
    """
    Ищет ВСЕ кабели с таким же ККС в базе (исключая исходный журнал).
    
    Args:
        sheet: лист Excel с данными
        cable_kks: ККС искомого кабеля
        source_journal: ККС исходного журнала (чтобы исключить его)
    
    Returns:
        list: список кортежей (журнал, источник, дата, строка_с_кабелем)
    """
    if not cable_kks:
        return []
    
    kks_col = get_column_index(sheet, 'ККС')
    if kks_col is None:
        return []
    
    found_cables = []
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            continue
        
        row_kks = str(row[kks_col - 1]).strip() if kks_col <= len(row) and row[kks_col - 1] else ''
        if row_kks != cable_kks:
            continue
        
        row_journal = str(row[0]).strip() if row[0] else ''
        if row_journal == source_journal:
            continue
        
        source, date = get_journal_info(sheet, row_journal, row_idx)
        found_cables.append((row_journal, source, date, row_idx))
    
    return found_cables


def compare_room_equip_pairs(original_from, original_to, response_from, response_to):
    """
    Сравнивает пары "Откуда" и "Куда" с учётом того, что пустые значения считаются "любыми".
    
    Логика:
    - ('a', 'b') == ('b', 'a') — порядок не важен
    - ('a', '') == ('a', 'b') — пустое равно любому
    - ('', 'b') == ('b', 'a') — пустое равно любому
    - ('', '') == ('a', '') — всегда True
    - ('', '') == ('a', 'b') — всегда True
    - ('a', 'b') != ('c', '') — не совпадают
    - ('a', '') != ('', 'b') — не совпадают
    
    Returns:
        bool: True если пары совпадают, False если нет
    """
    # Приводим к строкам и убираем пробелы
    of = str(original_from).strip() if original_from else ''
    ot = str(original_to).strip() if original_to else ''
    rf = str(response_from).strip() if response_from else ''
    rt = str(response_to).strip() if response_to else ''
    
    # Если оба значения пустые — считаем, что совпадают с любыми
    if not of and not ot:
        return True
    if not rf and not rt:
        return True
    
    # Собираем множества (порядок не важен), убираем пустые значения
    original_set = {x for x in (of, ot) if x}
    response_set = {x for x in (rf, rt) if x}
    
    # Если оба множества пустые — совпадают
    if not original_set and not response_set:
        return True
    
    # Если одно множество пустое, а другое нет — считаем, что совпадают
    if not original_set or not response_set:
        return True
    
    # Проверяем, что все значения из меньшего множества присутствуют в большем
    # (с учётом того, что пустые значения считаются "любыми")
    if len(original_set) <= len(response_set):
        return original_set.issubset(response_set)
    else:
        return response_set.issubset(original_set)


def extract_trace_elements(trace_str):
    """Извлекает из строки трассы все элементы, соответствующие regular_cable_tray"""
    import re
    from config import regular_cable_tray
    elements = set()
    # Убираем запятую в конце, если есть
    trace_clean = trace_str.rstrip(',').strip()
    # Разбиваем на слова (по пробелам и запятым)
    words = re.split(r'[\s,;]+', trace_clean)
    for word in words:
        word = word.strip()
        if word and regular_cable_tray.search(word):
            elements.add(word)
    return elements


def combine_traces(from_room, to_room, traces_list):
    """
    Объединяет трассы из нескольких кабелей в одну с учётом направления.
    

    def combine_traces должна получать на вход строку помещение откуда первого из объединяемых кабелей; строку помещение куда первого из объединяемых кабелей; список трасс группы объединяемых кабелей.
    1. Получить из строк ККС помещений подстроки  ККС здания - это regular_KKS_building. В итоге у нас есть ККС здания откуда и ККС здания куда. это начало и конец общей объединенной трассы
    2. каждую из объединяемых трасс разбить на слова и привести к списку трасс согласно regular_cable_tray: trace_list_1, trace_list_2 и т.д. по числу кабелей в групппе. порядок внутри списков важен. 
    3. списки выстроить в последовательность по следующей логике:
    есть ККС_здания_откуда , ККС_здания_куда
    есть trace_list_1, trace_list_2, ..., trace_list_n

    Если ККС_здания_откуда содержится как подстрока хотя бы в одном элементе trace_list и не содержится в остальных trace_list, то этот trace_list первый
    Если ККС_здания_откуда содержится как подстрока во всех элементах trace_list , то этот trace_list первый

    Если trace_list всего два и определено, который первый, оставшийся автоматически последний, последовательность выстроена.

    Если ККС_здания_куда содержится как подстрока хотя бы в одном элементе trace_list и не содержится в остальных trace_list, то этот trace_list последний
    Если ККС_здания_куда содержится как подстрока во всех элементах  trace_list, то этот trace_list последний

    Если trace_list всего два и определено, который последний, оставшийся автоматически первый, последовательность выстроена.

    Если trace_list больше двух и определено, который первый и который последний, оставшиеся расположить между ними, последовательность выстроена.

    Если ККС_здания_откуда или ККС_здания_куда отсутствуют или не находятся в элементах trace_list, последовательность выстроить по порядку.

    4. В первом trace_list если первый элемент не содержит как подстроку ККС_здания_откуда, то в этом trace_list изменить последовательность элементов на обратную
    5.  В последнем trace_list если последний элемент не содержит как подстроку ККС_здания_откуда, то в этом trace_list изменить последовательность элементов на обратную
    6. создать итоговую трассу - строку, соединяя элементы из trace_list в выстроенной последовательности через запятую. Между элементами разных trace_list  добавлять знак '+'
    7. полученную строку выдать

    Args:
        from_room: ККС помещения "Откуда" (первого кабеля в группе)
        to_room: ККС помещения "Куда" (первого кабеля в группе)
        traces_list: список строк с трассами (по одной на каждый кабель в группе)
    
    Returns:
        str: объединённая трасса
    """
    import re
    from config import regular_cable_tray, regular_KKS_building
    
    # 1. Получаем ККС здания откуда и куда
    from_building = ''
    to_building = ''
    
    if from_room:
        match = regular_KKS_building.search(str(from_room))
        if match:
            from_building = match.group()
    
    if to_room:
        match = regular_KKS_building.search(str(to_room))
        if match:
            to_building = match.group()
    
    # 2. Разбиваем каждую трассу на список элементов по regular_cable_tray
    trace_lists = []
    for trace in traces_list:
        if not trace:
            trace_lists.append([])
            continue
        elements = []
        # Убираем запятую в конце
        trace_clean = str(trace).rstrip(',').strip()
        # Разбиваем на слова (по пробелам и запятым)
        words = re.split(r'[\s,;]+', trace_clean)
        for word in words:
            word = word.strip()
            if word and regular_cable_tray.search(word):
                elements.append(word)
        trace_lists.append(elements)
    
    # Удаляем пустые списки
    trace_lists = [lst for lst in trace_lists if lst]
    
    if not trace_lists:
        return ''
    
    if len(trace_lists) == 1:
        # Если только одна трасса — возвращаем как есть
        return ', '.join(trace_lists[0])
    
    # 3. Определяем порядок списков
    # Функция для проверки, содержит ли список заданное здание
    def list_contains_building(lst, building):
        if not building:
            return False
        for item in lst:
            if building in item:
                return True
        return False
    
    # Функция для подсчёта, сколько списков содержат здание
    def count_lists_with_building(lists, building):
        if not building:
            return 0
        count = 0
        for lst in lists:
            if list_contains_building(lst, building):
                count += 1
        return count
    
    first_idx = None
    last_idx = None
    
    # Определяем первый список (по from_building)
    if from_building:
        from_count = count_lists_with_building(trace_lists, from_building)
        if from_count == 1:
            # Если здание только в одном списке — он первый
            for i, lst in enumerate(trace_lists):
                if list_contains_building(lst, from_building):
                    first_idx = i
                    break
        elif from_count == len(trace_lists):
            # Если здание во всех списках — берём первый список как первый
            first_idx = 0
    
    # Определяем последний список (по to_building)
    if to_building:
        to_count = count_lists_with_building(trace_lists, to_building)
        if to_count == 1:
            # Если здание только в одном списке — он последний
            for i, lst in enumerate(trace_lists):
                if list_contains_building(lst, to_building):
                    last_idx = i
                    break
        elif to_count == len(trace_lists):
            # Если здание во всех списках — берём последний список как последний
            last_idx = len(trace_lists) - 1
    
    # Упорядочиваем списки
    ordered_lists = []
    remaining_indices = list(range(len(trace_lists)))
    
    # Если определён первый индекс
    if first_idx is not None:
        ordered_lists.append(trace_lists[first_idx])
        remaining_indices.remove(first_idx)
    
    # Если определён последний индекс (и он не совпадает с первым)
    if last_idx is not None and last_idx != first_idx:
        # Если последний ещё не добавлен
        if last_idx in remaining_indices:
            # Добавляем все остальные, а последний — в конце
            for i in remaining_indices:
                if i != last_idx:
                    ordered_lists.append(trace_lists[i])
            ordered_lists.append(trace_lists[last_idx])
        else:
            # Если последний уже добавлен как первый (не должно случиться, но на всякий случай)
            for i in remaining_indices:
                ordered_lists.append(trace_lists[i])
    else:
        # Если последний не определён или совпадает с первым
        for i in remaining_indices:
            ordered_lists.append(trace_lists[i])
    
    # Если порядок не был определён (нет from_building и to_building) — оставляем как есть
    if first_idx is None and last_idx is None:
        ordered_lists = trace_lists.copy()
    
    # 4. В первом списке: если первый элемент не содержит from_building — разворачиваем
    if ordered_lists and from_building:
        first_list = ordered_lists[0]
        if first_list:
            # Проверяем первый элемент списка
            if not (from_building in first_list[0]):
                ordered_lists[0] = first_list[::-1]
    
    # 5. В последнем списке: если последний элемент не содержит to_building — разворачиваем
    if ordered_lists and to_building and len(ordered_lists) > 1:
        last_list = ordered_lists[-1]
        if last_list:
            # Проверяем последний элемент списка
            if not (to_building in last_list[-1]):
                ordered_lists[-1] = last_list[::-1]
    
    # 6. Собираем итоговую трассу
    result_parts = []
    for i, lst in enumerate(ordered_lists):
        if i > 0:
            # Между разными списками добавляем ' + '
            result_parts.append(' + ')
        result_parts.append(', '.join(lst))
    
    return ''.join(result_parts)


def create_worksheet(wb_out, results, sheet, journal_kks):
    """
    Создаёт лист 'В работу' с группами объединяемых кабелей.
    """
    # Создаём лист
    ws = wb_out.create_sheet("В работу")
    
    # Заголовки
    headers = [
        '№ п/п',
        'ККС кабеля',
        'Журнал',
        'Марка',
        'Сечение',
        'Группа',
        'Откуда помещение',
        'Откуда оборудование',
        'Куда помещение',
        'Куда оборудование',
        'Длина',
        'Трасса'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['L'].width = 50
    ws.freeze_panes = 'A2'
    
    # ========== ПОСТРОЕНИЕ ИНДЕКСА ==========
    # Один раз проходим по листу и строим словарь { (kks, journal): row }
    kks_col = get_column_index(sheet, 'ККС')
    row_index = {}
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        row_journal = str(row[0]).strip() if row[0] else ''
        row_kks = str(row[kks_col - 1]).strip() if kks_col and kks_col <= len(row) and row[kks_col - 1] else ''
        if row_kks and row_journal:
            key = (row_kks, row_journal)
            if key not in row_index:
                row_index[key] = row
    # ======================================
    
    # Получаем список колонок для извлечения данных из исходного листа
    col_map = {}
    for header in headers[2:]:
        col_idx = get_column_index(sheet, header)
        if col_idx:
            col_map[header] = col_idx
    
    row_idx = 2
    
    # Группируем результаты по ККС
    groups = {}
    for result in results:
        kks = result['kks']
        if not kks:
            continue
        if not result['response'] or result['response'] == '-':
            continue
        if kks not in groups:
            groups[kks] = {
                'num': result['num'],
                'source': result['source'],
                'cables': []
            }
        groups[kks]['cables'].append({
            'journal': result['response'],
            'response': result['response'],
            'note': result['note'],
            'problems': result['problems']
        })
    
    # Обрабатываем каждую группу
    for kks, group_data in groups.items():
        cables_data = []
        
        # Кабель из исходного журнала — ищем по индексу
        source_key = (kks, journal_kks)
        if source_key in row_index:
            cables_data.append({
                'journal': journal_kks,
                'is_source': True,
                'row': row_index[source_key]
            })
        
        # Ответные кабели — ищем по индексу
        for resp_info in group_data['cables']:
            resp_journal = resp_info['response'].split(' (')[0] if resp_info['response'] and resp_info['response'] != '-' else ''
            if not resp_journal:
                continue
            resp_key = (kks, resp_journal)
            if resp_key in row_index:
                cables_data.append({
                    'journal': resp_journal,
                    'is_source': False,
                    'row': row_index[resp_key]
                })
        
        # Записываем все кабели группы
        for idx, cable in enumerate(cables_data):
            row = cable['row']
            
            if idx == 0:
                ws.cell(row=row_idx, column=1, value=group_data['num'])
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                ws.cell(row=row_idx, column=7).fill = yellow_fill
                ws.cell(row=row_idx, column=9).fill = yellow_fill
            else:
                ws.cell(row=row_idx, column=1, value='')
            
            ws.cell(row=row_idx, column=2, value=kks)
            ws.cell(row=row_idx, column=3, value=cable['journal'])
            
            col_mapping = {
                4: 'Марка',
                5: 'Сечение',
                6: 'Группа',
                7: 'Откуда помещение',
                8: 'Откуда оборудование',
                9: 'Куда помещение',
                10: 'Куда оборудование',
                11: 'Длина',
                12: 'Трасса'
            }
            
            for col_num, header_name in col_mapping.items():
                if header_name in col_map:
                    val = row[col_map[header_name] - 1] if col_map[header_name] <= len(row) else ''
                    ws.cell(row=row_idx, column=col_num, value=str(val).strip() if val else '')
            
            row_idx += 1
        
        # Строка "Объединенный"
        if cables_data:
            first_row = cables_data[0]['row']
            ws.cell(row=row_idx, column=2, value=kks)
            ws.cell(row=row_idx, column=3, value="Объединенный")
            
            col_mapping_merged = {
                4: 'Марка',
                5: 'Сечение',
                6: 'Группа',
                7: 'Откуда помещение',
                8: 'Откуда оборудование',
                9: 'Куда помещение',
                10: 'Куда оборудование'
            }
            
            for col_num, header_name in col_mapping_merged.items():
                if header_name in col_map:
                    val = first_row[col_map[header_name] - 1] if col_map[header_name] <= len(first_row) else ''
                    ws.cell(row=row_idx, column=col_num, value=str(val).strip() if val else '')
            
            total_length = 0.0
            for cable in cables_data:
                row = cable['row']
                if 'Длина' in col_map:
                    length_val = row[col_map['Длина'] - 1] if col_map['Длина'] <= len(row) else 0
                    try:
                        total_length += float(str(length_val).replace(',', '.'))
                    except:
                        pass
            ws.cell(row=row_idx, column=11, value=round(total_length, 2) if total_length > 0 else '')
            
            traces = []
            from_room = ''
            to_room = ''
            for idx, cable in enumerate(cables_data):
                row = cable['row']
                if 'Трасса' in col_map:
                    trace_val = row[col_map['Трасса'] - 1] if col_map['Трасса'] <= len(row) else ''
                    if trace_val:
                        traces.append(str(trace_val).strip())
                if idx == 0:
                    if 'Откуда помещение' in col_map:
                        from_room = row[col_map['Откуда помещение'] - 1] if col_map['Откуда помещение'] <= len(row) else ''
                    if 'Куда помещение' in col_map:
                        to_room = row[col_map['Куда помещение'] - 1] if col_map['Куда помещение'] <= len(row) else ''
            
            merged_trace = combine_traces(from_room, to_room, traces) if traces else ''
            ws.cell(row=row_idx, column=12, value=merged_trace)
            row_idx += 1
        
        row_idx += 1
    
    return ws


def process_journal(excel_path, journal_kks, output_path):
    """
    Основная функция обработки журнала.
    
    Args:
        excel_path: путь к базе данных
        journal_kks: ККС журнала для обработки
        output_path: путь для сохранения результата
    """
    print(f"\n{'='*60}")
    print("ПОИСК ОТВЕТНЫХ ЧАСТЕЙ КАБЕЛЕЙ")
    print(f"{'='*60}")
    print(f"Загрузка: {excel_path}")
    print(f"Обработка журнала: {journal_kks}")
    
    wb = load_workbook(excel_path, data_only=True)
    sheet = wb.active
    
    # Находим индексы нужных колонок
    kks_col = get_column_index(sheet, 'ККС')
    status_col = get_column_index(sheet, 'Статус объединения')
    req_col = get_column_index(sheet, 'Требования к объединению')
    note_col = get_column_index(sheet, 'Примечание')
    
    if kks_col is None:
        raise ValueError("Колонка 'ККС' не найдена")
    
    # ========== ПОСТРОЕНИЕ ИНДЕКСА (ОДИН РАЗ) ==========
    row_index = {}
    for r in sheet.iter_rows(min_row=2, values_only=True):
        if not r:
            continue
        row_journal = str(r[0]).strip() if r[0] else ''
        row_kks = str(r[kks_col - 1]).strip() if kks_col and kks_col <= len(r) and r[kks_col - 1] else ''
        if row_kks and row_journal:
            key = (row_kks, row_journal)
            if key not in row_index:
                row_index[key] = r
    # ================================================
    
    # Создаём выходной файл
    wb_out = Workbook()
    sheet_out = wb_out.active
    sheet_out.title = "Отчёт"
    
    for head in HEADERS:
        col = head[0] + 1
        sheet_out.cell(row=1, column=col, value=head[1])
        sheet_out.column_dimensions[get_column_letter(col)].width = head[2]
    sheet_out.freeze_panes = 'A2'
    
    # Собираем все кабели из указанного журнала
    cables = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            continue
        row_journal = str(row[0]).strip() if row[0] else ''
        if row_journal == journal_kks:
            cables.append((row_idx, row))
    
    print(f"Найдено кабелей в журнале: {len(cables)}")
    
    # Обрабатываем каждый кабель
    results = []
    problem_count = 0
    result_counter = 0
    
    for row_idx, row in cables:
        # Проверяем условия: статус пуст И требования не пусты
        status = str(row[status_col - 1]).strip() if status_col and status_col <= len(row) and row[status_col - 1] else ''
        requirements = str(row[req_col - 1]).strip() if req_col and req_col <= len(row) and row[req_col - 1] else ''
        
        if status or not requirements:
            continue
        
        cable_kks = str(row[kks_col - 1]).strip() if kks_col <= len(row) and row[kks_col - 1] else ''
        if not cable_kks:
            continue

        # Получаем информацию об исходном журнале
        src_source, src_date = get_journal_info(sheet, journal_kks, row_idx)
        src_info = f"{journal_kks} ({src_source}, {src_date})"
        note = str(row[note_col - 1]).strip() if note_col and note_col <= len(row) and row[note_col - 1] else ''
        
        # Ищем ВСЕ ответные кабели во всей базе
        found_cables = find_all_matching_cables(sheet, cable_kks, journal_kks)
        
        if found_cables:
            # Сортируем по журналу для стабильности
            found_cables.sort(key=lambda x: x[0])
            
            for idx, (resp_journal, resp_source, resp_date, resp_row) in enumerate(found_cables):
                cable_problems = []
                
                # Получаем строку ответного кабеля ИЗ ИНДЕКСА (без сканирования)
                resp_key = (cable_kks, resp_journal)
                resp_row_data = row_index.get(resp_key)
                
                if resp_row_data:
                    # Проверка марки + сечения
                    mark_col = get_column_index(sheet, 'Марка')
                    section_col = get_column_index(sheet, 'Сечение')
                    original_mark = str(row[mark_col - 1]).strip() if mark_col and mark_col <= len(row) and row[mark_col - 1] else ''
                    original_section = str(row[section_col - 1]).strip() if section_col and section_col <= len(row) and row[section_col - 1] else ''
                    original_full = f"{original_mark} {original_section}".strip() if original_mark or original_section else ''
                    # Собираем марку + сечение для ответного кабеля
                    response_mark = str(resp_row_data[mark_col - 1]).strip() if mark_col and mark_col <= len(resp_row_data) and resp_row_data[mark_col - 1] else ''
                    response_section = str(resp_row_data[section_col - 1]).strip() if section_col and section_col <= len(resp_row_data) and resp_row_data[section_col - 1] else ''
                    response_full = f"{response_mark} {response_section}".strip() if response_mark or response_section else ''
                    
                    if original_full and response_full and original_full != response_full:
                        cable_problems.append(f"Марка не совпадает: {original_full} ≠ {response_full}")
                    
                    # Проверка помещений
                    from_room_col = get_column_index(sheet, 'Откуда помещение')
                    to_room_col = get_column_index(sheet, 'Куда помещение')
                    
                    if from_room_col and to_room_col:
                        original_from = str(row[from_room_col - 1]).strip() if row[from_room_col - 1] else ''
                        original_to = str(row[to_room_col - 1]).strip() if row[to_room_col - 1] else ''
                        response_from = str(resp_row_data[from_room_col - 1]).strip() if resp_row_data[from_room_col - 1] else ''
                        response_to = str(resp_row_data[to_room_col - 1]).strip() if resp_row_data[to_room_col - 1] else ''
                        
                        if not compare_room_equip_pairs(original_from, original_to, response_from, response_to):
                            original_str = f"{original_from} ↔ {original_to}" if original_from or original_to else '(пусто)'
                            response_str = f"{response_from} ↔ {response_to}" if response_from or response_to else '(пусто)'
                            cable_problems.append(f"Помещения не совпадают: {original_str} ≠ {response_str}")
                    
                    # Проверка оборудования
                    from_equip_col = get_column_index(sheet, 'Откуда оборудование')
                    to_equip_col = get_column_index(sheet, 'Куда оборудование')
                    
                    if from_equip_col and to_equip_col:
                        original_from_eq = str(row[from_equip_col - 1]).strip() if row[from_equip_col - 1] else ''
                        original_to_eq = str(row[to_equip_col - 1]).strip() if row[to_equip_col - 1] else ''
                        response_from_eq = str(resp_row_data[from_equip_col - 1]).strip() if resp_row_data[from_equip_col - 1] else ''
                        response_to_eq = str(resp_row_data[to_equip_col - 1]).strip() if resp_row_data[to_equip_col - 1] else ''
                        
                        if not compare_room_equip_pairs(original_from_eq, original_to_eq, response_from_eq, response_to_eq):
                            original_str = f"{original_from_eq} ↔ {original_to_eq}" if original_from_eq or original_to_eq else '(пусто)'
                            response_str = f"{response_from_eq} ↔ {response_to_eq}" if response_from_eq or response_to_eq else '(пусто)'
                            cable_problems.append(f"Оборудование не совпадает: {original_str} ≠ {response_str}")
                    
                    # Проверка вложенности трасс (разбиваем на отдельные элементы)
                    trace_col = get_column_index(sheet, 'Трасса')
                    if trace_col and trace_col <= len(resp_row_data):
                        original_trace = str(row[trace_col - 1]).strip() if row[trace_col - 1] else ''
                        response_trace = str(resp_row_data[trace_col - 1]).strip() if resp_row_data[trace_col - 1] else ''
                        if original_trace and response_trace:
                            original_elements = extract_trace_elements(original_trace)
                            response_elements = extract_trace_elements(response_trace)
                            if original_elements and response_elements:
                                if original_elements.issubset(response_elements) or original_elements == response_elements:
                                    cable_problems.append("Трасса исходного кабеля содержится в трассе ответного")
                                elif response_elements.issubset(original_elements):
                                    cable_problems.append("Трасса ответного кабеля содержится в трассе исходного")
                    
                    # Проверка группы
                    group_col = get_column_index(sheet, 'Группа')
                    if group_col and group_col <= len(resp_row_data):
                        original_group = str(row[group_col - 1]).strip() if row[group_col - 1] else ''
                        response_group = str(resp_row_data[group_col - 1]).strip() if resp_row_data[group_col - 1] else ''
                        if original_group and response_group and original_group != response_group:
                            cable_problems.append(f"Группа не совпадает: {original_group} ≠ {response_group}")
                
                if not cable_problems:
                    cable_problems.append("Проблем не обнаружено")
                
                if idx == 0:
                    result_counter += 1
                    results.append({
                        'num': result_counter,
                        'kks': cable_kks,
                        'source': src_info,
                        'response': f"{resp_journal} ({resp_source}, {resp_date})",
                        'note': note if note else '-',
                        'problems': '; '.join(cable_problems)
                    })
                else:
                    # Для остальных — num пустой, но kks и source заполняем!
                    results.append({
                        'num': '',
                        'kks': cable_kks,
                        'source': src_info,
                        'response': f"{resp_journal} ({resp_source}, {resp_date})",
                        'note': '',
                        'problems': '; '.join(cable_problems)
                    })
        else:
            result_counter += 1
            results.append({
                'num': result_counter,
                'kks': cable_kks,
                'source': src_info,
                'response': '-',
                'note': note if note else '-',
                'problems': 'Ответная часть не найдена в базе'
            })
            problem_count += 1
    

    # Записываем результаты в Excel
    row_idx = 2
    for result in results:
        sheet_out.cell(row=row_idx, column=1, value=result['num'])
        sheet_out.cell(row=row_idx, column=2, value=result['kks'])
        sheet_out.cell(row=row_idx, column=3, value=result['source'])
        sheet_out.cell(row=row_idx, column=4, value=result['response'])
        sheet_out.cell(row=row_idx, column=5, value=result['note'])
        sheet_out.cell(row=row_idx, column=6, value=result['problems'])
        
        row_idx += 1
    

    # ========== СОЗДАНИЕ ЛИСТА "РАБОЧИЙ СТОЛ" ==========
    print("\n📊 Создание листа 'Рабочий стол'...")
    create_worksheet(wb_out, results, sheet, journal_kks)
    # ====================================================
    
    # Применяем автофильтр ко всем листам
    for ws in wb_out.worksheets:
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

    # Сохраняем
    wb_out.save(output_path)

    print(f"\n{'='*60}")
    print("СТАТИСТИКА")
    print(f"{'='*60}")
    print(f"   Всего кабелей в журнале: {len(cables)}")
    print(f"   Обработано: {len(results)}")
    print(f"   С проблемами: {problem_count}")
    print(f"   Без проблем: {len(results) - problem_count}")
    print(f"\n✅ Сохранено: {output_path}")
    print(f"{'='*60}")
    
    return results


def process_multiple_journals(excel_path, journal_kks_list, output_dir=None):
    """
    Обрабатывает несколько журналов за один раз.
    
    Args:
        excel_path: путь к базе данных
        journal_kks_list: список ККС журналов для обработки
        output_dir: папка для сохранения отчётов (если None — рядом с базой)
    """
    if not journal_kks_list:
        print("❌ Список журналов пуст.")
        return
    
    print(f"\n{'='*60}")
    print("ПОИСК ОТВЕТНЫХ ЧАСТЕЙ ДЛЯ НЕСКОЛЬКИХ ЖУРНАЛОВ")
    print(f"{'='*60}")
    print(f"Загрузка: {excel_path}")
    print(f"Журналов для обработки: {len(journal_kks_list)}")
    print("-" * 60)
    
    # Создаём папку для отчётов
    db_path = Path(excel_path)
    if output_dir is None:
        reports_dir = db_path.parent / "Отчёты"
    else:
        reports_dir = Path(output_dir)
    reports_dir.mkdir(exist_ok=True)
    
    total_cables = 0
    total_results = 0
    total_problems = 0
    processed = []
    failed = []
    
    for i, journal_kks in enumerate(journal_kks_list, 1):
        journal_kks = journal_kks.strip()
        if not journal_kks:
            continue
        
        print(f"\n[{i}/{len(journal_kks_list)}] Обработка: {journal_kks}")
        print("-" * 40)
        
        try:
            # Формируем имя выходного файла
            safe_journal = journal_kks.replace('\\', '_').replace('/', '_').replace(':', '_')
            output_filename = f"Отчет по журналу {safe_journal}.xlsx"
            output_path = reports_dir / output_filename
            
            # Обрабатываем журнал
            results = process_journal(excel_path, journal_kks, str(output_path))
            
            # Собираем статистику
            total_cables += len(results)
            problem_count = sum(1 for r in results if r['problems'] and 'не найдена' not in r['problems'] and r['problems'] != 'Проблем не обнаружено')
            total_problems += problem_count
            total_results += len(results)
            
            processed.append({
                'journal': journal_kks,
                'cables': len(results),
                'problems': problem_count,
                'output': str(output_path)
            })
            
            print(f"  ✅ Готово: {len(results)} кабелей, {problem_count} с проблемами")
            
        except Exception as e:
            print(f"  ❌ Ошибка: {e}")
            failed.append({'journal': journal_kks, 'error': str(e)})
    
    # Выводим сводку
    print(f"\n{'='*60}")
    print("СВОДКА ПО ОБРАБОТКЕ")
    print(f"{'='*60}")
    print(f"  Всего журналов: {len(journal_kks_list)}")
    print(f"  Успешно обработано: {len(processed)}")
    print(f"  С ошибками: {len(failed)}")
    print(f"  Всего кабелей: {total_results}")
    print(f"  Всего проблем: {total_problems}")
    
    if processed:
        print("\n  Успешно обработанные журналы:")
        for p in processed:
            print(f"    {p['journal']} — {p['cables']} кабелей ({p['problems']} проблем)")
    
    if failed:
        print("\n  Ошибки:")
        for f in failed:
            print(f"    {f['journal']} — {f['error']}")
    
    print(f"\n✅ Отчёты сохранены в: {reports_dir}")
    print(f"{'='*60}")
    
    return processed, failed


# ========== ТОЧКА ВХОДА ===============
if __name__ == "__main__":
    print("\n" + "="*60)
    print("ПОИСК ОТВЕТНЫХ ЧАСТЕЙ КАБЕЛЕЙ")
    print("="*60)
    print("\nВыберите режим работы:")
    print("  1 — Один журнал")
    print("  2 — Несколько журналов (список)")
    print()
    
    mode = input("Ваш выбор (1 или 2): ").strip()
    
    # Ввод пути к базе
    excel_path = input("Введите путь к файлу базы данных (Cable base ver.*.xlsx): ").strip()
    excel_path = excel_path.strip('"').strip("'")
    
    if not excel_path:
        print("❌ Путь не указан. Программа завершена.")
        sys.exit(1)
    
    if not Path(excel_path).exists():
        print(f"❌ Файл не найден: {excel_path}")
        sys.exit(1)
    
    if mode == '2':
        # Режим нескольких журналов
        print("\nВведите ККС журналов для поиска (по одному на строку).")
        print("Для завершения ввода оставьте пустую строку и нажмите Enter.")
        print("-" * 40)
        
        journals = []
        while True:
            line = input().strip()
            if not line:
                break
            journals.append(line)
        
        if not journals:
            print("❌ Не введено ни одного журнала. Программа завершена.")
            sys.exit(1)
        
        output_dir = input("\nВведите папку для сохранения отчётов (Enter для папки 'Отчёты' рядом с базой): ").strip()
        output_dir = output_dir.strip('"').strip("'") if output_dir else None
        
        try:
            process_multiple_journals(excel_path, journals, output_dir)
            print("\n🏁 Готово!")
        except Exception as e:
            print(f"\n❌ Ошибка: {e}")
            sys.exit(1)
    
    else:
        # Режим одного журнала
        journal_kks = input("Введите ККС журнала для поиска ответных частей: ").strip()
        journal_kks = journal_kks.strip('"').strip("'")
        
        if not journal_kks:
            print("❌ ККС журнала не введён. Программа завершена.")
            sys.exit(1)
        
        output_path = input("Введите путь для сохранения результата (имя_файла.xlsx): ").strip()
        output_path = output_path.strip('"').strip("'")
        
        if not output_path:
            print("❌ Путь для сохранения не указан. Программа завершена.")
            sys.exit(1)
        
        if not output_path.endswith('.xlsx'):
            output_path += '.xlsx'
        
        try:
            process_journal(excel_path, journal_kks, output_path)
            print("\n🏁 Готово!")
        except Exception as e:
            print(f"\n❌ Ошибка: {e}")
            sys.exit(1)