
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import pyodbc 
from docx import Document
from docx.enum.text import WD_ALIGN_PARAGRAPH # Стили
from docx.enum.section import WD_SECTION, WD_ORIENT # Стили
from docx.shared import Mm, Cm # Для установки размеров в мм
from pathlib import Path
from datetime import date

from collections import defaultdict

from JournalMaster.file_operations import convert_doc_to_docx
from Libraries.defWords import cleanCyrFromLat

from openpyxl.utils import get_column_letter


def unity_finder(dir_in, dir_out):
    """
    Ищу в базе совпадения кабелей

    Args:
        dir_in (_type_): _description_
        dir_out (_type_): _description_
    """

    b_name = 'Cable base ver.'

    current_version = current_version_finder(dir_in, b_name) #  текущая версия базы, та, которую мы ДЕЛАЕМ сейчас. Но поскольку мы не создаем новую базу, а читаем ту, что сохранена последней, то:
    if current_version > 1:
        current_version -= 1

    print(f'сейчас текущая версия для открытия {current_version}')
    rb = load_workbook(dir_in + '/' + b_name + str(current_version) + '.xlsx', data_only=True) # нашел и открыл последнюю версию базы
    sheetRead = rb.active


    # создаю словарь из списков
    # defaultdict — это специальная версия обычного словаря (dict) из модуля collections, которая автоматически создаёт значение по умолчанию для отсутствующего ключа при обращении к нему.
    cables = defaultdict(list)  # list() создаёт пустой список

    for rowRead in range (2, sheetRead.max_row + 1):
        current_row = [] # список содержимого ячеек одной строки из базы 
        for colRead in range(1, sheetRead.max_column + 1):
            current_row.append(sheetRead.cell(row = rowRead, column = colRead).value)
        
        cables[current_row[2]].append([current_row[0], current_row[7], current_row[8]]) # добавляет словарь --  ККС (current_row[2]) : [журнал 0, длина 7, трасса 8]


    print(f'В текущей версии {current_version} имеем:')
    # вывожу список кабелей, у которых в базе есть повторяющиеся ККС
    for key, value in cables.items():
        if len(value) > 1:
            js = []
            for v in value:
                js.append(v[0])
            print(f"{key}: {js}")


    # wb = openpyxl.Workbook()
    # sheetWrite = wb.active
    # sheetWrite.title = "База данных вер. " + str(current_version)
